import constants
from constants import Constants
import openpyxl
import tkinter as tk
import RequestHandler

# Custom error classes. I wanted these so that when errors were triggered, the main code line could understand what
# error was triggered and record it accordingly.
class VariantAPINoItem(Exception):
    pass

class VariantAPITooMany(Exception):
    pass

class VariantAPINoPrices(Exception):
    pass

# General functions for use across all classes within main
def getDateDash():
    import datetime
    return datetime.datetime.now().strftime("%d-%m-%Y")


def getDateSlash():
    import datetime
    return datetime.datetime.now().strftime("%d/%m/%Y")

# Function for mapping country codes that come out of the API to their respective country names.
def countryCaseMapper():

    constants.Constants.countryCodeRef = {}
    for country in constants.Constants.countryCodeExchange:
        constants.Constants.countryCodeRef[country[0].upper()] = country[0]

    return None


# Class VariantRef for containing and manipulating data concerning BC item codes from input sheet and their properties.
class VariantRef:

    def __init__(self, dictionary):

        self.itemDict = dictionary
        self.priceDict = {}

        # If the API is active, either:
        # 1. Pull the prices from the items API (used when not closing off). This will then either use the full price,
        # markdown price or promotional price depending on the user's choice and multiplier columns existance.
        # 2. Pull the prices from the prices API (used when closing off). This will then use the price type specified
        # in the constants file and generate a file precise to whats in BC.

        if Constants.useAPI:
            if Constants.closeOff:
                self.pullExchangePriceDictionary()
            else:
                self.populateOpenPrices()
                self.buildExchangePriceDictionary()
        else:
            self.buildExchangePriceDictionary()

        self.startDate = {}
        self.endDate = {}
        self.populateDates()

    # Using exchange rates from constants, build a dictionary of exchange prices for each country.
    def buildExchangePriceDictionary(self):

        import math
        for country in Constants.countryCodeExchange:

            # Please note, DO NOT CHANGE (round(float(. https://stackoverflow.com/questions/455612/limiting-floats-to-two-decimal-points
            # Floats are not precise, so the float from the input spreadsheet must be rounded to 2 decimal places.

            self.priceDict[country[0]] = math.ceil((round(float(self.itemDict[Constants.inputPriceColName]), 2) * country[1]))

        self.itemDict["Exchange Prices"] = self.priceDict

    # Go to BC and get the prices for the item code in the input sheet. This is used when closing off.
    def pullExchangePriceDictionary(self):

        priceRequest = RequestHandler.bcGetRequest("REDACTED"
                                                   + self.itemDict[Constants.inputIdentifierColName] + "'")

        if priceRequest is None or priceRequest == []:
            raise VariantAPINoPrices("No prices found in BC, item code: " + self.itemDict[Constants.inputIdentifierColName])

        else:
            for price in priceRequest["value"]:

                # Format customer e.g SHOPIFY CANADA becomes Canada
                customer = Constants.countryCodeRef[(str(price["customerNo"])).replace(Constants.customerNumberPrefix.upper(), "")]

                # Get price from api response dependent on which price the user wants. There is likely a cleaner way
                # to do this via string splitting but this is more readable.

                if constants.Constants.priceType == "Full Price":
                    self.priceDict[customer] = price["fullPrice"]

                elif constants.Constants.priceType == "Markdown Price":
                    self.priceDict[customer] = price["markdownPrice"]

                elif constants.Constants.priceType == "Promotional Price":
                    self.priceDict[customer] = price["promotionalPrice"]

            self.itemDict["Exchange Prices"] = self.priceDict

    # If the user has not selected closing off, it is assumed that the rrp will be used as the full price in the items
    # API. A price will be generated here based on the user's choice of price type.
    def populateOpenPrices(self):

        # Call the BC API to get the general price for the item under the "rrp" field
        itemRequest = RequestHandler.bcGetRequest("REDACTED'"
                                               + self.itemDict[Constants.inputIdentifierColName] + "'")

        # If the BC API returns nothing, raise an error
        if itemRequest is None or itemRequest == []:
            raise VariantAPINoItem("Item not found in BC, item code: " + self.itemDict[Constants.inputIdentifierColName])

        # If the BC API returns more than one item, raise an error
        elif len(itemRequest["value"]) > 1:
            raise VariantAPITooMany("Multiple items found in BC, item code: " + self.itemDict[Constants.inputIdentifierColName])

        # If the setting is "Full Price", nothing needs to be done as rrp should be Shopify UK price. On the other hand,
        # if the setting is "Promotional" or "Markdown" then the rrp needs to be recalculated

        # To calculate MK or Promo price, the rrp is multiplied by 1 - the discount percentage. This is then rounded
        # indiscriminately.

        else:
            if constants.Constants.priceType != "Full Price":
                self.itemDict[Constants.inputPriceColName] = \
                    round(itemRequest["value"][0]["rrp"] * (1 - float(self.itemDict[Constants.inputPriceColName])))

            else:
                self.itemDict[Constants.inputPriceColName] = itemRequest["value"][0]["rrp"]

    # Give the item a start and end date based on the user's initial inputs.
    def populateDates(self):

        if Constants.startDateCustomization == 2:
            self.itemDict["Starting Date"] = Constants.customStartDate
        elif Constants.startDateCustomization == 3:
            self.itemDict["Starting Date"] = self.itemDict[Constants.inputStartDateColName]
        else:
            self.itemDict["Starting Date"] = getDateSlash()

        if Constants.customEndDateBool:
            self.itemDict["Ending Date"] = Constants.customEndDate

        else:
            self.itemDict["Ending Date"] = None

    # Method used to return all item properties.
    def getVar(self):
        return self.itemDict

# Manager class for managing workbook operations, namely the opening, saving, reading, editing and closing of workbooks.
# TODO: A separate workbook class may be a good implementation, namely for having more than 2 workbooks open at once.
class WorkbookManager:

    def loadInputWorkbook(self, inputBookName=Constants.inputBookName):

        self.inputWorkbook = openpyxl.load_workbook(inputBookName)

    def loadOutputWorkbook(self):

        self.outputWorkbook = openpyxl.Workbook()

    def getInputSheet(self):

        return self.inputWorkbook.active

    def closeInputWorkbook(self):

        self.inputWorkbook.close()

    def closeOutputWorkbook(self):

        self.outputWorkbook.close()

    # Column finder takes a parameter that is a list of column names to search for. This is by default the column names
    # in Constants i.e the user provided ones however the testing also requires an input for the output column names,
    # hence this is parameterized.
    def columnFinder(self, identifiers=Constants.inputColumnIdentifiers):

        inputSheet = self.getInputSheet()

        columnPositions = {}
        # Search the radius specified in constants to look for the column names also specified in cosntants.
        for row in inputSheet.iter_rows(max_row=Constants.searchRadius,
                                           max_col=Constants.searchRadius):
            for cell in row:
                if cell.value in identifiers:
                    # If the column name is found, add the row to the return array.
                    columnPositions[cell.value] = [cell.row, cell.column, cell.value]

        if len(columnPositions) != len(identifiers):
            raise Exception(str("Not all column identifiers found in input sheet. Identifiers not found: " + str(identifiers)))

        return columnPositions

    # Does what it says on the tin. Returns the minimum and maximum columns in the input sheet.
    def getMinAndMaxColumns(self, columnPositions):

        keys = list(columnPositions.keys())
        minColumn = columnPositions[keys[0]][0]
        maxColumn = columnPositions[keys[0]][1]

        for column in columnPositions:
            if columnPositions[column][0] < minColumn:
                minColumn = columnPositions[column][0]
            if columnPositions[column][1] > maxColumn:
                maxColumn = columnPositions[column][1]

        return minColumn, maxColumn

    # This function is used to get the input data from the input sheet. It is parameterized to allow for the tests
    # to use it as well as the main code (i.e columns to get will change with the conversion test as an output
    # file can be passed back through here).
    # Single process is used to identify whether the sheet is an input or output. If it is an input, the data is
    # stored in a dictionary but with just one price, the price in the input sheet. If it is an output, the data is
    # stored in a dictionary with multiple prices, the prices in the output sheet.
    # Count none is used to identify whether the function should count the number of consecutive rows with no data.

    def getInputData(self, columnsToGet=Constants.inputColumnIdentifiers,
                     identifierColumnName=Constants.inputIdentifierColName, singleProcess=True,
                     countNone=True):

        ####################################

        def noneCountHandler(noneCount):

            if noneCount > Constants.noneCountTolerance:
                return noneCount, True
            else:
                noneCount += 1
                return noneCount, False

        def addSingleColumnToIdentifierDictionary(columnData, inputSheet, dict, identifierList):

            # Column data is in format [row, column, column name]

            # Note, I thought about many ways to allow this function to account for inputting output files as there is
            # a different number of rows to account for. This is a hard problem to solve and needs alot of consideration -
            # I have decided to leave it for now as it is not a priority.
            # The difficulty is that input files need to allow for columns in any y position while output files need to
            # allow for rows of identical data. These concepts are hard to merge.

            noneCount = 0
            breakLoop = False
            for row in inputSheet.iter_rows(min_row=columnData[0] + 1, max_row=(len(identifierList) + 1),
                                            min_col=columnData[1], max_col=columnData[1]):
                # -2 is used to account for the fact that the identifier list index starts at 0,
                # whereas the data starts at row 2  (row 1 is the column names)

                if breakLoop:
                    break

                elif row[0].value is None:
                    noneCount, breakLoop = noneCountHandler(noneCount)
                    noneList.append(row[0].row)
                    continue

                dict[identifierList[row[0].row-2]][columnData[2]] = row[0].value

            return dict

        inputSheet = self.getInputSheet()
        columnPositions = self.columnFinder(columnsToGet)

        # Column positions variable is in format {column name: [row, column, column name]}

        dataDictFinal = {}
        identifierDetails = columnPositions[identifierColumnName]

        # identifierDetails is in format [row, column, column name]
        noneCount = 0
        noneList = []

        # This loop builds the identifier dictionary, appending the identifier along with its row number.
        # Maybe the identifier dictionary can be combined with the single column builder? Look into this

        breakLoop = False
        for row in inputSheet.iter_rows(min_row=identifierDetails[0] + 1, max_row=inputSheet.max_row,
                                                     min_col=identifierDetails[1], max_col=identifierDetails[1]):

            # If the number of consecutive rows with no data exceeds the tolerance, break the loop. This is to account
            # for the fact that the input sheet may have empty rows at the bottom.

            if breakLoop:
                break

            elif row[0].value is None:
                noneCount, breakLoop = noneCountHandler(noneCount)
                noneList.append([row[0].row, row[0].column])
                continue

            else:
                dataDictFinal[row[0].value] = {row[0].value: row[0].row}
                dataDictFinal[row[0].value][identifierColumnName] = row[0].value

        identifierList = list(dataDictFinal.keys())

        if singleProcess:
            for column in columnPositions:

                if column == Constants.inputIdentifierColName:
                    continue

                dataDictFinal = addSingleColumnToIdentifierDictionary(columnPositions[column], inputSheet, dataDictFinal,
                                                                identifierList)


        else:
            # Min and max returned as a tuple in format (min, max)
            minAndMax = self.getMinAndMaxColumns(columnPositions)#

            generalIdentifierColumn = columnPositions[Constants.outputGeneralIdentifier][1]
            shopifyCodeColumn = columnPositions[Constants.outputShopifyCodeName][1]
            priceColumn = columnPositions[Constants.outputPriceUsed][1]

            for column in columnPositions:

                noneCount = 0

                # Note that dataDictFinal is stored like ["identifier":{"identifier":row number, "column name": data}]

                if column == Constants.outputGeneralIdentifier:
                    continue

                elif column == Constants.outputPriceUsed:
                    continue

                elif column == Constants.outputShopifyCodeName:

                    breakLoop = False
                    for row in inputSheet.iter_rows(min_row=columnPositions[column][0] + 1, max_row=inputSheet.max_row,
                                                    min_col=minAndMax[0], max_col=minAndMax[1]):

                        # If the column in the row contains a shopify code, add it to the dictionary along with the
                        # found price

                        # A reminder column positions is in format {column name: [row, column, column name]}

                        identifierVal = row[generalIdentifierColumn - 1].value

                        if breakLoop:
                            break

                        elif identifierVal is None:
                            noneCount, breakLoop = noneCountHandler(noneCount)
                            noneList.append([row[0].row, row[0].column])
                            continue

                        dataDictFinal[identifierVal][row[shopifyCodeColumn - 1].value] = row[priceColumn - 1].value

                else:

                    breakLoop = False
                    for row in inputSheet.iter_rows(min_row=columnPositions[column][0] + 1, max_row=inputSheet.max_row,
                                                    min_col=minAndMax[0], max_col=minAndMax[1]):

                        identifierVal = row[generalIdentifierColumn - 1].value

                        if breakLoop:
                            break

                        elif identifierVal is None:
                            noneCount, breakLoop = noneCountHandler(noneCount)
                            noneList.append([row[0].row, row[0].column])
                            continue

                        dataDictFinal[identifierVal][column] = row[columnPositions[column][1] - 1].value

        return dataDictFinal
    def getOutputFile(self, varList):

        self.loadOutputWorkbook()
        
        outputSheet = self.outputWorkbook.active
        outputSheet.append(Constants.outputColumnNames)

        # The code below lacks modularity. This is formatted simply for the BC input format, and therefore will
        # need changing if that format changes

        for var in varList:

            varData = var.getVar()
            prices = varData["Exchange Prices"]

            for x, code in enumerate(prices):
                tempRow = [varData[Constants.inputIdentifierColName], Constants.customerNumberPrefix + str(code),
                           varData["Starting Date"], varData["Ending Date"], Constants.priceType,
                           prices[code]]
                outputSheet.append(tempRow)

        outputBookName = str(getDateDash()) + Constants.outputBookSuffix
        self.outputWorkbook.save(outputBookName)
        constants.Constants.outputBookName = outputBookName
        self.outputWorkbook.close()

# Deals with data integrity and formatting.
class DataFormatter:

    def __init__(self):

        self.noneCounter = 0

    def embeddedListToDictionary(self, inputList):
        returnList = []
        keys = inputList[0]

        for row in inputList[1:]:
            returnList.append(dict(zip(keys, row)))

        return returnList

    def cellFormatter(self, inputList):

        import datetime
        for x, list in enumerate(inputList):
            for y, cell in enumerate(list):
                if type(cell) == str:
                    cell = cell.strip()
                elif type(cell) == int:
                    cell = str(cell)
                elif type(cell) == float:
                    cell = str(cell)
                elif type(cell) == datetime.datetime:
                    cell = datetime.datetime.strftime(cell, "%d/%m/%Y")
                else:
                    raise Exception("Unknown cell type found in input data, cell type was: " + str(type(cell)))
                inputList[x][y] = cell

        return inputList

    def cellFormatterDict(self, inputDict):

        import datetime

        for key in inputDict:
            for property in inputDict[key]:
                if type(inputDict[key][property]) == str:
                    inputDict[key][property] = inputDict[key][property].strip()
                elif type(inputDict[key][property]) == int:
                    inputDict[key][property] = str(inputDict[key][property])
                elif type(inputDict[key][property]) == float:
                    inputDict[key][property] = str(inputDict[key][property])
                elif type(inputDict[key][property]) == datetime.datetime:
                    inputDict[key][property] = datetime.datetime.strftime(inputDict[key][property], "%d/%m/%Y")
                else:
                    raise Exception("Unknown cell type found in input data, cell type was: " +
                                    str(type(inputDict[key][property])))

        return inputDict
    def removeNoneRows(self, inputData):

        popList = []
        itemCodeLevel = inputData.keys()
        for itemCode in itemCodeLevel:
            for itemPropertyKey in inputData[itemCode]:

                # TODO: Temporary fix for ending date
                if itemPropertyKey == "Ending Date":
                    continue

                if inputData[itemCode][itemPropertyKey] is None:
                    popList.append(itemCode)
                    self.noneCounter+=1
                    break

        for itemCode in popList:
            popper = inputData.pop(itemCode)

        return inputData


    def convertCurrency(self, val, currency):

        import math
        for array in Constants.countryCodeExchange:
            if array[0] == currency:
                return math.ceil((round(float(val), 2) * array[1]))



class Controller:

    def __init__(self):
        pass

    def run(self, progressBarCallback):

        try:

            progressBarCallback(25, "Initiating workbook managers 1/3")
            workbookManager = WorkbookManager()
            dataFormatter = DataFormatter()
            countryCaseMapper()

            progressBarCallback(25, "Initiating loading input workbook data 1/3")
            workbookManager.loadInputWorkbook(Constants.originalBookLocation)

            # If 1, search by column names expected to be in the first row of the input file, if 2, search for columns
            # via radius method.
            progressBarCallback(25, "Reading input sheet 1/3")
            inputData = workbookManager.getInputData(Constants.inputColumnIdentifiers,
                                                     Constants.inputIdentifierColName, True, True)

            progressBarCallback(5, "Formatting cells from book 1/3")
            inputData = dataFormatter.cellFormatterDict(inputData)

            progressBarCallback(5, "Removing none rows from book 1/3")
            inputData = dataFormatter.removeNoneRows(inputData)

            varObjectList = []

            progressBarCallback(15, "Creating currency conversions for each item 1/3", True)

            increment = 100 / len(inputData)

            for dictionary in inputData:

                try:
                    progressBarCallback(increment, "Processing item: " + str(dictionary) + ". 2/3")
                    varObjectList.append(VariantRef(inputData[dictionary]))

                except Exception as e:
                    print("Found issue with line: " + str(dictionary) + "\n" + str(e.with_traceback(None)))

            progressBarCallback(0, "Finished processing items. 3/3", True)
            progressBarCallback(50, "Creating output file from book 3/3")
            workbookManager.getOutputFile(varObjectList)

            progressBarCallback(50, "Program Finished! Book used was: " + Constants.originalBookName)
            return None

        except Exception as e:
            return e


## Ideas i had for future development
# TODO: Add more data integrity checks and formatters
# TODO: Error handling system
# Sheet Splitting
# Configurable column translations
